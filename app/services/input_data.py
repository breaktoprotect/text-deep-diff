from typing import List, Dict, Union
from pathlib import Path
from openpyxl import load_workbook


class UnsupportedFileTypeError(Exception):
    pass


def _validate_excel(file_path: Union[str, Path]) -> None:
    if not str(file_path).lower().endswith((".xlsx", ".xls")):
        raise UnsupportedFileTypeError(f"Unsupported file type: {file_path}")


def list_sheets(file_path: Union[str, Path]) -> List[str]:
    """
    Returns list of sheet names in an Excel file.
    """
    _validate_excel(file_path)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    return wb.sheetnames


def list_columns(file_path: Union[str, Path], sheet_name: str) -> List[str]:
    """
    Returns the column headers of the given Excel sheet.
    """
    _validate_excel(file_path)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return [col for col in header_row if col is not None]


def extract_data(
    file_path: Union[str, Path],
    sheet_name: str,
    columns: List[str],
) -> List[Dict[str, Union[str, float]]]:
    """
    Extracts rows as list of dicts from the specified Excel sheet using selected columns.
    """
    _validate_excel(file_path)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb[sheet_name]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    col_idx_map = {name: idx for idx, name in enumerate(header) if name in columns}

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {
            col: row[idx]
            for col, idx in col_idx_map.items()
            if idx < len(row) and row[idx] is not None
        }
        if any(row_data.values()):
            data.append(row_data)
    return data


def construct_sentences(
    file_path: Union[str, Path],
    sheet_name: str,
    columns: List[str],
    separator: str = " ",
) -> List[str]:
    """
    Constructs sentences by joining selected column values for each row.
    """
    rows = extract_data(file_path, sheet_name, columns)
    sentences = []
    for row in rows:
        parts = [
            str(row[col]).strip()
            for col in columns
            if col in row and row[col] is not None
        ]
        sentence = separator.join(parts)
        if sentence:
            sentences.append(sentence)
    return sentences
